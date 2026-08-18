# ─────────────────────────────────────────────────────────────────────────────
# pedidos.py  —  Gestión de pedidos (JSON pendientes + Excel confirmados)
# ─────────────────────────────────────────────────────────────────────────────

import json
import os
import io
import datetime
import openpyxl
from filelock import FileLock
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from PIL import Image as PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

from config import PEDIDOS_XLSX, PEDIDOS_JSON

# ── Lock para evitar race conditions ─────────────────────────────────────────
_LOCK_PATH = PEDIDOS_JSON + ".lock"

def _cargar_json():
    if os.path.exists(PEDIDOS_JSON):
        with open(PEDIDOS_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def _guardar_json(data):
    os.makedirs(os.path.dirname(PEDIDOS_JSON), exist_ok=True)
    with open(PEDIDOS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def siguiente_ref():
    """Devuelve la siguiente referencia de pedido (ej. P-0042). Thread-safe."""
    with FileLock(_LOCK_PATH):
        data  = _cargar_json()
        nuevo = data.get("_contador", 0) + 1
        data["_contador"] = nuevo
        _guardar_json(data)
    return f"P-{nuevo:04d}"

def guardar_pedido_pendiente(ref: str, pedido: dict):
    """Guarda un pedido en estado PENDIENTE_PAGO. Thread-safe."""
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
        data[ref] = {**pedido, "ref": ref, "estado": "PENDIENTE_PAGO",
                     "fecha": datetime.datetime.now().isoformat()}
        _guardar_json(data)

def obtener_pedido(ref: str):
    with FileLock(_LOCK_PATH):
        return _cargar_json().get(ref)

def marcar_confirmado(ref: str):
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
        if ref in data:
            data[ref]["estado"] = "CONFIRMADO"
            data[ref]["fecha_confirmacion"] = datetime.datetime.now().isoformat()
            _guardar_json(data)

def buscar_pedido_pendiente_por_usuario(cliente_id: int):
    """Devuelve el pedido más reciente en PENDIENTE_PAGO de ese usuario, o None."""
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
    candidatos = [
        v for k, v in data.items()
        if k != "_contador"
        and v.get("cliente_id") == cliente_id
        and v.get("estado") == "PENDIENTE_PAGO"
    ]
    if not candidatos:
        return None
    return sorted(candidatos, key=lambda x: x.get("fecha", ""), reverse=True)[0]

def marcar_comprobante_enviado(ref: str):
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
        if ref in data:
            data[ref]["estado"] = "COMPROBANTE_ENVIADO"
            _guardar_json(data)

def marcar_rechazado(ref: str):
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
        if ref in data:
            data[ref]["estado"] = "RECHAZADO"
            _guardar_json(data)

def pedidos_pendientes_pago():
    with FileLock(_LOCK_PATH):
        data = _cargar_json()
    return {k: v for k, v in data.items()
            if k != "_contador" and v.get("estado") == "PENDIENTE_PAGO"}

# ── Excel de pedidos confirmados ──────────────────────────────────────────────

HEADERS = [
    "Ref", "Fecha", "Cliente (nombre)", "Cliente (Telegram)",
    "Equipo", "Temporada", "Producto", "Foto", "Ref. proveedor", "Público",
    "Versión", "Talla", "Nombre dorsal", "Número dorsal",
    "Parche", "Cantidad", "Precio unit. (€)", "Descuento (%)", "Total (€)", "Estado",
]
_COL_FOTO = HEADERS.index("Foto") + 1   # columna 1-based de la foto
_FOTO_PX  = 60                           # tamaño miniatura en píxeles

thin  = Side(style="thin", color="BBBBBB")
brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

def _init_excel():
    """Crea el Excel de pedidos si no existe."""
    os.makedirs(os.path.dirname(PEDIDOS_XLSX), exist_ok=True)
    if os.path.exists(PEDIDOS_XLSX):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PEDIDOS"

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    c = ws["A1"]
    c.value = "REGISTRO DE PEDIDOS — Bot Camisetas"
    c.font  = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="1A237E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Cabeceras
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font  = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor="283593")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[2].height = 30

    # Anchos (col Foto = 9 caracteres ≈ 60px)
    anchos = [10,18,20,20,10,10,30,9,30,8,10,10,18,10,8,8,14,12,10,14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    wb.save(PEDIDOS_XLSX)


def _insertar_foto_excel(ws, fila, foto_path):
    """Incrusta una miniatura de la foto en la celda de la columna Foto."""
    if not _PIL_OK or not foto_path or not os.path.exists(foto_path):
        return
    try:
        pil_img = PILImage.open(foto_path).convert("RGB")
        pil_img.thumbnail((_FOTO_PX, _FOTO_PX), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width  = _FOTO_PX
        xl_img.height = _FOTO_PX
        col_letra = get_column_letter(_COL_FOTO)
        xl_img.anchor = f"{col_letra}{fila}"
        ws.add_image(xl_img)
        ws.row_dimensions[fila].height = 50   # ≈ 67px para que quepa la foto
    except Exception:
        pass  # Si falla, la fila queda sin foto (nunca rompe el guardado)


def _escribir_fila_excel(ws, fila, pedido, nombre_dorsal, numero_dorsal, cantidad, precio_unit, item=None):
    """
    Escribe una fila en el Excel.

    Si `item` se proporciona, los campos de producto (equipo, temporada,
    producto_nombre, nombre_proveedor, publico, talla, parche, foto) se leen de él.
    Los campos de cliente siempre vienen de `pedido`.
    """
    descuento_pct = pedido.get("descuento_pct", 0)
    total_bruto   = precio_unit * cantidad
    total         = round(total_bruto * (1 - descuento_pct / 100), 2) if descuento_pct else total_bruto

    cliente_tg = (
        f"@{pedido.get('cliente_username', '')}"
        if pedido.get("cliente_username")
        else str(pedido.get("cliente_id", ""))
    )

    # Fuente de campos de producto: item (nuevo formato) o pedido (legado)
    src = item if item is not None else pedido
    foto_path = src.get("foto", "") or pedido.get("foto", "")

    # Valores de columnas — "Foto" va en _COL_FOTO (celda vacía, la imagen se superpone)
    valores = [
        pedido.get("ref", ""),                                         # 1  Ref
        datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),            # 2  Fecha
        pedido.get("cliente_nombre", ""),                              # 3  Cliente nombre
        cliente_tg,                                                    # 4  Cliente TG
        src.get("equipo",           pedido.get("equipo", "")),         # 5  Equipo
        src.get("temporada",        pedido.get("temporada", "")),      # 6  Temporada
        src.get("producto_nombre",  pedido.get("producto_nombre", "")),# 7  Producto
        "",                                                            # 8  Foto (vacío, imagen encima)
        src.get("nombre_proveedor", pedido.get("nombre_proveedor", "")),# 9 Ref. proveedor
        src.get("publico",          pedido.get("publico", "")),        # 10 Público
        pedido.get("version", ""),                                     # 11 Versión
        src.get("talla",            pedido.get("talla", "")),          # 12 Talla
        nombre_dorsal or "—",                                          # 13 Nombre dorsal
        numero_dorsal or "—",                                          # 14 Número dorsal
        src.get("parche",           pedido.get("parche", "No")),       # 15 Parche
        cantidad,                                                      # 16 Cantidad
        precio_unit,                                                   # 17 Precio unit.
        f"{descuento_pct}%" if descuento_pct else "—",                # 18 Descuento
        total,                                                         # 19 Total
        "CONFIRMADO",                                                  # 20 Estado
    ]
    fill_fila = PatternFill("solid", fgColor="E8F5E9")
    for ci, val in enumerate(valores, 1):
        c = ws.cell(row=fila, column=ci, value=val)
        c.font      = Font(name="Arial", size=9)
        c.fill      = fill_fila
        c.border    = brd
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[fila].height = 20

    # Incrustar miniatura de la foto
    _insertar_foto_excel(ws, fila, foto_path)

    return total


def registrar_pedido_excel(pedido: dict):
    """Añade una o varias filas al Excel con el pedido confirmado.

    Soporta tanto el nuevo formato (con ``items`` array) como el formato
    plano antiguo (compatibilidad hacia atrás).
    """
    _init_excel()
    wb = openpyxl.load_workbook(PEDIDOS_XLSX)
    ws = wb.active

    items = pedido.get("items")

    if items:
        # ── Nuevo formato: iterar sobre los items del carrito ──────────────
        for item in items:
            precio_unit       = item.get("precio_unit", 0)
            cantidad          = item.get("cantidad", 1)
            personalizaciones = item.get("personalizaciones", [])

            if len(personalizaciones) > 1:
                # Una fila por personalización (cada camiseta individualizada)
                for p in personalizaciones:
                    fila = ws.max_row + 1
                    _escribir_fila_excel(
                        ws, fila, pedido,
                        p["nombre"], p["numero"],
                        1, precio_unit, item=item
                    )
            else:
                # Una fila con la cantidad total
                fila   = ws.max_row + 1
                nombre = personalizaciones[0]["nombre"] if personalizaciones else ""
                numero = personalizaciones[0]["numero"]  if personalizaciones else ""
                _escribir_fila_excel(
                    ws, fila, pedido,
                    nombre, numero,
                    cantidad, precio_unit, item=item
                )
    else:
        # ── Formato plano antiguo (backward compatibility) ─────────────────
        precio_unit       = pedido.get("precio_unit", 0)
        cantidad          = pedido.get("cantidad", 1)
        personalizaciones = pedido.get("personalizaciones", [])

        if len(personalizaciones) > 1:
            for p in personalizaciones:
                fila = ws.max_row + 1
                _escribir_fila_excel(
                    ws, fila, pedido, p["nombre"], p["numero"], 1, precio_unit
                )
        else:
            fila   = ws.max_row + 1
            nombre = personalizaciones[0]["nombre"] if personalizaciones else pedido.get("nombre_dorsal", "")
            numero = personalizaciones[0]["numero"]  if personalizaciones else pedido.get("numero_dorsal", "")
            _escribir_fila_excel(
                ws, fila, pedido, nombre, numero, cantidad, precio_unit
            )

    wb.save(PEDIDOS_XLSX)
